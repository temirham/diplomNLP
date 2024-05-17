import csv
import openpyxl
from pymystem3 import Mystem
from classificator import predict_detail_class_SVM, predict_detail_class_NB, predict_detail_class_DT
import re



def is_number(token):
    pattern = r"^\d+"  # Шаблон для поиска начала строки, за которым следует одно или более цифр
    match = re.match(pattern, token)
    if match:
        return int(match.group())  # Возвращаем числовое значение из токена
    else:
        return None  # Если токен не содержит числа, возвращаем None



def process_detail_text(text: str, predicted_class: str):
    # Создаем новую книгу Excel
    workbook = openpyxl.Workbook()
    # Выбираем активный лист
    sheet = workbook.active
    pattern = r"\d+"  # Шаблон для поиска одной или более цифр
    mystem_analyzer = Mystem()
    lemmatized_text = mystem_analyzer.lemmatize(text.lower())
    words_to_remove = ["равна", "равный", "равно", predicted_class, 'составлять', 'из', 'себя']
    cleaned_text = [token for token in lemmatized_text if token.strip() and token not in words_to_remove]
    cleaned_text.extend([""] * 4)
    print(cleaned_text)

    data_dict = {
        "двойной_ряд_шаровой_подшипник": {
            "внутренний диаметр": "",
            "внешний диаметр": "",
            "обозначение модели": "",
            "толщина": "",
        },
        "гайка": {
            "обозначение резьба": "",
            "ширина по плоскость": "",
            "диаметр головки": "",
            "толщина": "",
            "диаметр отверстие": ""
        },
        "профиль": {
            "тип": "",
            "длина балки": "",
            "радиус скругления балки": "",
            "ширина полки": "",
            "толщина полотна": "",
            "высота балки": "",
            "толщина фланца": ""
        },
        "радиальный подшипник": {
            "обозначение модели": "",
            "тип защиты подшипника": "",
            "радиус скругления": "",
            "толщина": "",
            "постфикс": "",
            "внешний диаметр": "",
            "внутренний диаметр": ""
        },
        "батарея": {
            "диаметр": "",
            "тип код": "",
            "высота": ""
        },
        'труба': {
            "внутренний диаметр": "",
            'наружный диаметр':'',
            'длина':''
        },
        'болт': {
            'диаметр резьбового вала': '',
            'длина вала без резьбы' : '',
            'длина головы' : '',
            'ширина по плоскостям' : '',
            'диаметр головки' : '',
            'длина без учета головы' : '',
            'Обозначение резьбы' : ''
        }
    }

    # # Предсказываем класс детали
    # predicted_class = predict_detail_class(text)
    # print("Предсказанный класс детали:", predicted_class)

    # Определяем форму детали
    data_dict[predicted_class]["форма"] = predicted_class

    # Итерируемся по токенам и заполняем словарь
    if predicted_class == "двойной_ряд_шаровой_подшипник":
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "обозначение" and cleaned_text[i + 1] == "модель":
                data_dict[predicted_class]["обозначение модели"] = cleaned_text[i+2]  # Получаем название детали
            elif cleaned_text[i] == "внутренний" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i+2]):
                data_dict[predicted_class]["внутренний диаметр"] = is_number(cleaned_text[i+2])  # Получаем внутренний диаметр
            elif cleaned_text[i] == "внешний" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i+2]):
                data_dict[predicted_class]["внешний диаметр"] = is_number(cleaned_text[i+2])  # Получаем внешний диаметр
            elif cleaned_text[i] == "толщина" and is_number(cleaned_text[i+1]):
                data_dict[predicted_class]["толщина"] = is_number(cleaned_text[i+1])  # Получаем толщину
    elif predicted_class == "гайка":
        # Условия для гайки
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "обозначение" and cleaned_text[i+1] == "резьба":
                data_dict[predicted_class]["обозначение резьба"] = cleaned_text[i+2]  # Получаем обозначение резьбы
            elif cleaned_text[i] == "ширина" and cleaned_text[i+1] == "по" and cleaned_text[i+2] == "плоскость" and is_number(cleaned_text[i + 3]):
                data_dict[predicted_class]["ширина по плоскость"] = is_number(cleaned_text[i + 3])  # Получаем ширину по плоскости
            elif cleaned_text[i] == "диаметр" and cleaned_text[i+1] == "головка" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["диаметр головки"] = is_number(cleaned_text[i + 2])  # Получаем диаметр головки
            elif cleaned_text[i] == "толщина" and is_number(cleaned_text[i + 1]):
                data_dict[predicted_class]["толщина"] = is_number(cleaned_text[i + 1])  # Получаем толщину
            elif cleaned_text[i] == "диаметр" and cleaned_text[i+1] == "отверстие" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["диаметр отверстие"] = is_number(cleaned_text[i + 2])  # Получаем диаметр отверстия
    elif predicted_class == "профиль":
        # Условия для профиля
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "тип":
                data_dict[predicted_class]["тип"] = cleaned_text[i+1]  # Получаем тип
            elif cleaned_text[i] == "длина" and cleaned_text[i+1] == "балка" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["длина балки"] = is_number(cleaned_text[i + 2])  # Получаем длину балок
            elif cleaned_text[i] ==  "радиус" and cleaned_text[i+1] == "скругление" and cleaned_text[i+2] == "балка" and is_number(cleaned_text[i + 3]):
                data_dict[predicted_class]["радиус скругления балки"] = is_number(cleaned_text[i + 3])  # Получаем радиус скругления балок
            elif cleaned_text[i] == "ширина" and cleaned_text[i+1] == "полка" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["ширина полки"] = is_number(cleaned_text[i + 2])  # Получаем ширину полки
            elif cleaned_text[i] == "толщина" and cleaned_text[i+1] == "полотно" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["толщина полотна"] = is_number(cleaned_text[i + 2])  # Получаем толщину полотна
            elif cleaned_text[i] == "высота" and cleaned_text[i+1] == "балка" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["высота балки"] = is_number(cleaned_text[i + 2])  # Получаем высоту балки
            elif cleaned_text[i] == "толщина" and cleaned_text[i+1] == "фланец" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["толщина фланца"] = is_number(cleaned_text[i + 2])  # Получаем толщину фланца
    elif predicted_class == "радиальный подшипник":
        # Условия для радиального подшипника
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "обозначение" and cleaned_text[i+1] == "модель":
                data_dict[predicted_class]["обозначение модели"] = cleaned_text[i+2]  # Получаем обозначение модели
            elif cleaned_text[i] == "тип" and cleaned_text[i+1] == "защита" and cleaned_text[i+2] == "подшипник":
                data_dict[predicted_class]["тип защиты подшипника"] = cleaned_text[i+3]  # Получаем тип защиты подшипника
            elif cleaned_text[i] == "радиус" and cleaned_text[i+1] == "скругление" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["радиус скругления"] =  is_number(cleaned_text[i + 2]) # Получаем радиус скругления
            elif cleaned_text[i] == "толщина" and cleaned_text[i+1] == "постфикс" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["толщина постфикс"] = is_number(cleaned_text[i + 2]) # Получаем толщину постфикса
            elif cleaned_text[i] == "внешний" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["внешний диаметр"] = is_number(cleaned_text[i + 2])  # Получаем внешний диаметр
            elif cleaned_text[i] == "внутренний" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["внутренний диаметр"] = is_number(cleaned_text[i + 2])  # Получаем внутренний диаметр
    elif predicted_class == "батарея":
        # Условия для батареи
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "диаметр" and is_number(cleaned_text[i + 1]):
                data_dict[predicted_class]["диаметр"] = is_number(cleaned_text[i + 1])  # Получаем диаметр батареи
            elif cleaned_text[i] == "тип" and cleaned_text[i+1] == "код":
                data_dict[predicted_class]["тип код"] = cleaned_text[i+2]  # Получаем тип кода
            elif cleaned_text[i] == "высота" and is_number(cleaned_text[i + 1]):
                data_dict[predicted_class]["высота"] = is_number(cleaned_text[i + 1])  # Получаем высоту батареи
    elif predicted_class == "труба":
    # Условия для трубы
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "внутренний" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["внутренний диаметр"] = is_number(cleaned_text[i + 2])  # Получаем внутренний диаметр
            elif cleaned_text[i] == "наружный" and cleaned_text[i+1] == "диаметр" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["наружный диаметр"] = is_number(cleaned_text[i + 2])  # Получаем наружный диаметр
            elif cleaned_text[i] == "длина" and is_number(cleaned_text[i + 1]):
                data_dict[predicted_class]["длина"] = is_number(cleaned_text[i + 1])  # Получаем длину
    elif predicted_class == "болт":
        # Условия для болта
        for i in range(len(cleaned_text) - 4):
            if cleaned_text[i] == "диаметр" and cleaned_text[i+1] == "резьбовый" and cleaned_text[i+2] == "вал" and is_number(cleaned_text[i + 3]):
                data_dict[predicted_class]["диаметр резьбового вала"] = is_number(cleaned_text[i + 3])  # Получаем диаметр резьбового вала
            elif cleaned_text[i] == "длина" and cleaned_text[i+1] == "вал" and cleaned_text[i+2] == "без" and cleaned_text[i+3] == "резьбы" and is_number(cleaned_text[i + 4]):
                data_dict[predicted_class]["длина вала без резьбы"] = is_number(cleaned_text[i + 4])  # Получаем длину вала без резьбы
            elif cleaned_text[i] == "длина" and cleaned_text[i+1] == "голова" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["длина головы"] = is_number(cleaned_text[i + 2])  # Получаем длину головы
            elif cleaned_text[i] == "ширина" and cleaned_text[i+1] == "по" and cleaned_text[i+2] == "плоскость" and is_number(cleaned_text[i + 3]):
                data_dict[predicted_class]["ширина по плоскостям"] = is_number(cleaned_text[i + 3])  # Получаем ширину по плоскостям
            elif cleaned_text[i] == "диаметр" and cleaned_text[i+1] == "головка" and is_number(cleaned_text[i + 2]):
                data_dict[predicted_class]["диаметр головки"] = is_number(cleaned_text[i + 2])  # Получаем диаметр головки
            elif cleaned_text[i] == "длина" and cleaned_text[i+1] == "без" and cleaned_text[i+2] == "учет" and cleaned_text[i+3] == "голова" and is_number(cleaned_text[i + 4]):
                data_dict[predicted_class]["длина без учета головы"] = is_number(cleaned_text[i + 4])  # Получаем длину без учета головы
            elif cleaned_text[i] == "обозначение" and cleaned_text[i+1] == "резьба":
                data_dict[predicted_class]["Обозначение резьбы"] = cleaned_text[i+2]  # Получаем обозначение резьбы
    return(data_dict[predicted_class])


def save_to_csv(file_path, data_dict):
    with open(file_path, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=data_dict.keys(), delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerow(data_dict)

def save_to_excel(file_path, data_dict):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    for key, value in data_dict.items():
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=value)
        row += 1
    workbook.save(file_path)


text = "Данная деталь представляет собой трубу с внутренним диаметром 40 мм, наружным диаметром 50 мм и длиной 1000 мм"
predict_class, time_of_predict = predict_detail_class_DT(text)
dict1 = process_detail_text(text, predict_class)
save_to_csv('technical_data.csv', dict1)